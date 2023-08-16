from openpyxl import Workbook, load_workbook
from flask import Flask, request, render_template, redirect

app = Flask(__name__)

def add_to_excel(name, age, comments):
    try:
        wb = load_workbook('person_data.xlsx')
        sheet = wb['Data']
    except:
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Data'
        sheet.append(['Name', 'Alter', 'Kommentare'])

    sheet.append([name, age, comments])
    wb.save('person_data.xlsx')

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        name = request.form['name']
        age = int(request.form['age'])
        comments = request.form['comments']

        if 18 <= age <= 40:
            add_to_excel(name, age, comments)

        return redirect('/')
    return render_template('form.html')

if __name__ == '__main__':
    app.run(debug=True)
